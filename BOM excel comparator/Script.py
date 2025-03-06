import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import re
from openpyxl.comments import Comment
import time
from utilities import *

start_time = time.time()

COLOR_MODIFIED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
COLOR_MODIFIED_section = PatternFill(start_color="e3ac36", end_color="D3D3D3", fill_type="solid")

font = Font(color='FFFFFF')

# old_path =  'Test_Data/BOM/bom_ref.xlsx'
# new_path = 'Test_Data/BOM/bom_new.xlsx'
# output = 'Test_Data/BOM/bom_output.xlsx'

def get_idx_racine(idx_next,df):
    if df.iloc[idx_next].Level == 0 :
        return [idx_next]
    pr_level = df.iloc[idx_next].Level
    pr_level = re.findall(r'\d+', str(pr_level))
    pr_level = pr_level[0]
    # print(pr_level)
    
    idx_levels = [idx_next]
    idx_next = idx_next + 1
    while idx_next < len(df.index):
      nv = df.iloc[idx_next].Level
      nbr = re.findall(r'\d+', str(nv))[0]
      # print(nbr)
      if int(nbr) > int(pr_level) :
        idx_levels.append(idx_next)
      else:
        break
      idx_next = idx_next + 1
    return idx_levels


def compare_files(old_path, new_path, output):

    df_old = pd.read_excel(old_path,engine='openpyxl')
    df_new = pd.read_excel(new_path,engine='openpyxl')

    df_old = df_old.fillna('NAN')
    df_new = df_new.fillna('NAN')
    df_old['Changed on'] = df_old['Changed on'].astype(str)
    df_new['Changed on'] = df_new['Changed on'].astype(str)

    diff_values = df_old.where(df_new.ne(df_old), other="SAME")
    idx_diff = diff_values.index[diff_values.ne("SAME").any(axis=1)]


    wb_new = load_workbook(new_path)
    ws_new = wb_new.active

    if "Summary" not in wb_new.sheetnames:
        ws_summary = wb_new.create_sheet("Summary")
    else:
        ws_summary = wb_new["Summary"]

    ws_summary.delete_rows(1, ws_summary.max_row)

    for idx,col in enumerate(df_old.columns):
        cell = f'A{idx+2}'
        ws_summary[cell] = col


    row_sum, col_sum = 1,2

    for row_idx in idx_diff :
        for col_idx in range(1, df_new.shape[1]):
            old_value = df_old.iloc[row_idx,col_idx]
            new_value = df_new.iloc[row_idx,col_idx]
            
            if pd.isna(old_value) and pd.isna(new_value):                
                continue
            elif old_value != new_value :
                print(f'old value {old_value} ,new value {new_value}')
                cell = ws_new.cell(row=row_idx+2, column=col_idx+1)
                cell.fill = COLOR_MODIFIED
                cell.font = font
                cell.comment = Comment(add_comment(old_value, new_value), "AutoComparer")

                
        section = get_idx_racine(row_idx, df_old)
        print(f'Section : {section}')
        for row in section:
            ws_summary.cell(row=row_sum, column=col_sum, value=f"Row {row+2}")
            for col in range(df_new.shape[1]):
                row_sum += 1
                print(f'row {row}, col {col} //// value {df_new.iloc[row,col]}')
                cell_summary = ws_summary.cell(row=row_sum, column=col_sum, value=df_new.iloc[row,col])
                
                if df_old.iloc[row,col] != df_new.iloc[row,col] :
                    cell_summary.fill = COLOR_MODIFIED
                    cell_summary.font = font
                    cell_summary.comment = Comment(add_comment(df_old.iloc[row,col], df_new.iloc[row,col]), "AutoComparer")
                    continue

                cell_summary.fill = COLOR_MODIFIED_section
                cell_summary.font = font
                cell_summary.comment = Comment(add_comment(df_old.iloc[row,col], df_new.iloc[row,col]), "AutoComparer")
                cell = ws_new.cell(row=row+2, column=col+1)
                cell.fill = COLOR_MODIFIED_section
                cell.font = font
            col_sum +=1
            row_sum = 1
        col_sum +=1
            
    
    Number_of_rows = df_old.shape[0]
    Number_of_columns = df_old.shape[1]
    Number_of_rows_modified = len(idx_diff)
    Number_of_rows_unchanged = Number_of_rows - Number_of_rows_modified
    
    metrics = {
    4 : Number_of_rows_unchanged, 
    6 : Number_of_rows_modified, 
    8 : Number_of_columns, 
    10 : Number_of_rows
    }
    
    if "Summary metrics" not in wb_new.sheetnames:
        ws_summary = wb_new.create_sheet("Summary metrics")
    else:
        ws_summary = wb_new["Summary metrics"]
        
    ws_summary.delete_rows(1, ws_summary.max_row)

    ws_summary.merge_cells("D1:G3")
    ws_summary["D1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary["D1"] = "Dataset Modification Summary"
    ws_summary["D1"].font = Font(bold=True)
    
    for i in range(4,11,2) :
        ws_summary.merge_cells(f"D{i}:E{i}")
        ws_summary[f"D{i}"] = "Number of rows unchanged"
        ws_summary[f"D{i}"].font = Font(bold=True)
        ws_summary[f"G{i}"] = metrics[i]
        
    
                
    wb_new.save(output)
    end_time = time.time()

    execution_time = end_time - start_time
    print(f"Temps d'exécution : {execution_time} secondes")
    print(f"Comparison complete. Differences highlighted in {output}.")

