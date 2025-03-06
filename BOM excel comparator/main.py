import tkinter as tk
from tkinter import filedialog, Label, Button
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment
from PIL import Image, ImageTk
import pandas as pd
import re
import time
import os
import sys
import random



start_time = time.time()


def resource_path(relative_path):
    try:
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)
    except Exception as e:
        print(f"Erreur lors du chargement de la ressource : {e}")
        return relative_path
logo_path = resource_path("yazaki_logo.png")

def generate_colors(num_colors):
    colors = []
    for _ in range(num_colors):
        color = "{:02x}{:02x}{:02x}".format(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
        colors.append(color)
    return colors



def add_comment(old_value, new_value):
    old_text = "(None)" if old_value == 'NAN' else str(old_value)
    new_text = "(None)" if new_value == 'NAN' else str(new_value)
    comment_text = f"Previous: {old_text}\nNew: {new_text}"
    # print(comment_text+'\n')
    return comment_text

COLOR_MODIFIED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
COLOR_MODIFIED_section = PatternFill(start_color="e3ac36", end_color="D3D3D3", fill_type="solid")

font = Font(color='FFFFFF')


def get_precd_idx(idx, df):
    # Extraction du niveau de départ directement
    level_acc = int(re.findall(r'\d+', str(df.loc[idx, 'Level']))[0])
    
    # Initialisation des listes de résultats
    idx_arr = [idx]
    levels = [level_acc]
    
    # Parcours à l'envers sans revenir en arrière pour des niveaux égaux
    for idx_prec in range(idx - 1, -1, -1):
        level_prec = int(re.findall(r'\d+', str(df.loc[idx_prec, 'Level']))[0])
        
        if level_prec < levels[-1]:
            idx_arr.append(idx_prec)
            levels.append(level_prec)
        elif level_prec == levels[-1]:
            continue  # Continue la boucle sans changer
        elif levels[-1] == 0:
            break  # On sort si le niveau atteint 0
    
    return idx_arr


def compare_files(old_path, new_path, output):

    df_old = pd.read_excel(old_path,engine='openpyxl')
    df_new = pd.read_excel(new_path,engine='openpyxl')

    df_old = df_old.fillna('NAN')
    df_new = df_new.fillna('NAN')
    df_old['Changed on'] = df_old['Changed on'].astype(str)
    df_new['Changed on'] = df_new['Changed on'].astype(str)

    diff_values = df_old.where(df_new.ne(df_old), other="SAME")
    idx_diff = diff_values.index[diff_values.ne("SAME").any(axis=1)]
    colors = generate_colors(len(idx_diff))

    wb_new = load_workbook(new_path)
    ws_new = wb_new.active

    if "Summary" not in wb_new.sheetnames:
        ws_summary = wb_new.create_sheet("Summary")
    else:
        ws_summary = wb_new["Summary"]

    ws_summary.delete_rows(1, ws_summary.max_row)

    for idx,col in enumerate(df_old.columns):
        cell = f'A{idx+2}'
        ws_summary[cell] = col


    row_sum, col_sum = 1,2
    idx_color = 0
    for row_idx in idx_diff :
        for col_idx in range(1, df_new.shape[1]):
            old_value = df_old.iloc[row_idx,col_idx]
            new_value = df_new.iloc[row_idx,col_idx]
            
            if pd.isna(old_value) and pd.isna(new_value):                
                continue
            elif old_value != new_value :
                print(f'old value {old_value} ,new value {new_value}')
                cell = ws_new.cell(row=row_idx+2, column=col_idx+1)
                cell.fill = COLOR_MODIFIED
                cell.font = font
                cell.comment = Comment(add_comment(old_value, new_value), "AutoComparer")

                
        section = get_precd_idx(row_idx, df_old)
        print(f'Section : {section}')
        for row in section:
            ws_summary.cell(row=row_sum, column=col_sum, value=f"Row {row+2}")
            for col in range(df_new.shape[1]):
                row_sum += 1
                print(f'row {row}, col {col} //// value {df_new.iloc[row,col]}')
                cell_summary = ws_summary.cell(row=row_sum, column=col_sum, value=df_new.iloc[row,col])
                
                if df_old.iloc[row,col] != df_new.iloc[row,col] :
                    cell_summary.fill = COLOR_MODIFIED
                    cell_summary.font = font
                    cell_summary.comment = Comment(add_comment(df_old.iloc[row,col], df_new.iloc[row,col]), "AutoComparer")
                    continue

                cell_summary.fill = PatternFill(start_color=colors[idx_color], end_color= colors[idx_color], fill_type="solid")
                cell_summary.font = font
                cell_summary.comment = Comment(add_comment(df_old.iloc[row,col], df_new.iloc[row,col]), "AutoComparer")
                cell = ws_new.cell(row=row+2, column=col+1)
                cell.fill = PatternFill(start_color=colors[idx_color], end_color= colors[idx_color], fill_type="solid")
                cell.font = font
            col_sum +=1
            row_sum = 1
        col_sum +=1
        idx_color += 1
            
    
    Number_of_rows = df_old.shape[0]
    Number_of_columns = df_old.shape[1]
    Number_of_rows_modified = len(idx_diff)
    Number_of_rows_unchanged = Number_of_rows - Number_of_rows_modified
    
    metrics = {
        4 : (Number_of_rows_unchanged,'Number of rows unchanged'), 
        6 : (Number_of_rows_modified,'Number of rows modified'), 
        8 : (Number_of_columns,'Number of columns'), 
        10 : (Number_of_rows, 'Number of rows')
    }
    
    if "Summary metrics" not in wb_new.sheetnames:
        ws_summary = wb_new.create_sheet("Summary metrics")
    else:
        ws_summary = wb_new["Summary metrics"]
        
    ws_summary.delete_rows(1, ws_summary.max_row)

    ws_summary.merge_cells("D1:G3")
    ws_summary["D1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary["D1"] = "Dataset Modification Summary"
    ws_summary["D1"].font = Font(bold=True)
    
    for i in range(4,11,2) :
        ws_summary.merge_cells(f"D{i}:E{i}")
        ws_summary[f"D{i}"] = metrics[i][1]
        ws_summary[f"D{i}"].font = Font(bold=True)
        ws_summary[f"G{i}"] = metrics[i][0]
        
    
                
    wb_new.save(output)
    end_time = time.time()

    execution_time = (end_time - start_time)/60
    print(f"Temps d'exécution : {execution_time} minutes")
    print(f"Comparison complete. Differences highlighted in {output}.")
    return execution_time


def compare_excels():
    if not ref_file_path or not new_file_path:
        status_label.config(text="Veuillez sélectionner les fichiers nécessaires.", fg="red")
        return

    status_text = f"Référence: {ref_file_path}\nNouveau: {new_file_path}"
    if output_dir:
        status_text += f"\nDossier de sortie: {output_dir}"

    status_label.config(text=status_text, fg="black")
    
    try:
        if ref_file_path and new_file_path and output_dir:
            status_label.config(text="Comparaison en cours...", fg="red")
            execution_time = compare_files(ref_file_path, new_file_path, f"{output_dir}/output.xlsx")
            if execution_time : 
                status_label.config(text=f"Temps d'exécution : {execution_time} minutes", fg="black")
                output_file = f"{output_dir}/output.xlsx"
                os.startfile(output_file)
        else:
            status_label.config(text="Les chemins des fichiers ou du dossier de sortie sont manquants.", fg="red")
    except Exception as e:
        status_label.config(text=f"Erreur lors de la comparaison: {str(e)}", fg="red")
    

from tkinter import messagebox

def select_ref_file():
    global ref_file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.XLSX")])
    
    if not file_path.endswith((".xlsx", ".xls","*.XLSX")):
        messagebox.showerror("Erreur", "Veuillez sélectionner un fichier Excel valide (*.xlsx, *.xls;*.XLSX)")
        return

    ref_file_path = file_path
    ref_label.config(text=f"Référence: {ref_file_path}")

def select_new_file():
    global new_file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    
    if not file_path.endswith((".xlsx", ".xls")):
        messagebox.showerror("Erreur", "Veuillez sélectionner un fichier Excel valide (*.xlsx, *.xls)")
        return

    new_file_path = file_path
    new_label.config(text=f"Nouveau: {new_file_path}")


def select_output_dir():
    global output_dir
    output_dir = filedialog.askdirectory()
    output_label.config(text=f"Dossier de sortie: {output_dir}")


root = tk.Tk()
root.title("Excel Comparator - Yazaki")
root.geometry("350x350")

# window_width = 300
# window_height = 300

# # Obtenir les dimensions de l'écran
# screen_width = root.winfo_screenwidth()
# screen_height = root.winfo_screenheight()

# # Calculer la position x et y
# x_position = (screen_width - window_width) // 2
# y_position = (screen_height - window_height) // 2

# # Appliquer la position centrée
# root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")

# logo_path = "yazaki_logo.png"  
try:
    img = Image.open(logo_path)
    img = img.resize((150, 50), Image.LANCZOS)
    logo = ImageTk.PhotoImage(img)
    root.logo = logo
    logo_label = Label(root, image=logo)
    logo_label.pack(pady=10)
except:
    logo_label = Label(root, text="[Logo Yazaki]", font=("Arial", 14, "bold"))
    logo_label.pack(pady=10)


ref_file_path = ""
new_file_path = ""
output_dir = ""

ref_button = Button(root, text="Sélectionner le fichier de référence", command=select_ref_file)
ref_button.pack(pady=5)
ref_label = Label(root, text="Référence: Non sélectionné", wraplength=400)
ref_label.pack()

new_button = Button(root, text="Sélectionner le fichier nouveau", command=select_new_file)
new_button.pack(pady=5)
new_label = Label(root, text="Nouveau: Non sélectionné", wraplength=400)
new_label.pack()

output_button = Button(root, text="Sélectionner le dossier de sortie", command=select_output_dir)
output_button.pack(pady=5)
output_label = Label(root, text="Dossier de sortie: Non sélectionné", wraplength=400)
output_label.pack()

compare_button = Button(root, text="Comparer", command=compare_excels)
compare_button.pack(pady=20)

status_label = Label(root, text="", fg="black")
status_label.pack()

tk.mainloop()