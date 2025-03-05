from openpyxl import   load_workbook
from openpyxl.styles import PatternFill
from google.colab import files
import pandas as pd

_ = files.upload()

df = pd.read_excel('Electricity.xlsx')
moy_cons = df['Consumption'].mean()
moy_prod = df['Production'].mean()


wb = load_workbook('Electricity.xlsx')
ws =wb.active

def test_func(colonne, moy_colonne):
    if colonne >= moy_colonne :
        return '00FF00'
    elif colonne == moy_colonne :
        return 'FFFF00'
    elif colonne <= moy_colonne :
        return 'FF0000'


for row in ws.iter_rows(min_row=1,max_row=ws.max_row, min_col=1, max_col = 10):
    DateTime, Consumption, Production, Nuclear, Wind, Hydroelectric, Oil, Gas, Coal, Solar, Biomass  = row
    color_cons = test_func(Consumption.value, moy_cons)
    color_prod = test_func(Production.value, moy_prod)
    ws.cell(row=Consumption.row,column=2).fill = PatternFill(start_color=color_cons, end_color=color_cons, fill_type='solid')
    ws.cell(row=Production.row,column=3).fill = PatternFill(start_color=color_prod, end_color=color_prod, fill_type='solid')
        
        				 
wb.save('Electricity_formatted.xlsx')
