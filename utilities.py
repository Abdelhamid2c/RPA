from openpyxl.comments import Comment
from openpyxl.styles import Font
import pandas as pd

def add_comment(old_value, new_value):
    old_text = "(None)" if old_value == 'NAN' else str(old_value)
    new_text = "(None)" if new_value == 'NAN' else str(new_value)
    comment_text = f"Previous: {old_text}\nNew: {new_text}"
    # print(comment_text+'\n')
    return comment_text

def color_line(row, ws, highlight_color, Status):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = highlight_color   
        cell.font = Font(color='FFFFFF')
    if Status.lower() == 'deleted':
        comment_text = f"Row deleted"
        ws.cell(row=row, column=1).comment = Comment(comment_text, "Author")
    return ws    
    
        